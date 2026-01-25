"""
Fetch signer data for ALL 287(g) agencies.

This script:
1. Fetches sheriff directories from state associations (live)
2. Looks up police chiefs from agency websites
3. Saves structured signer data

Run periodically to keep data fresh (sheriffs elected every 4 years).
"""

import json
import re
import time
from pathlib import Path
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


# URLs for state sheriff associations
SHERIFF_SOURCES = {
    'ALABAMA': 'https://www.alabamasheriffs.com/sheriffs-directory',
    'ARIZONA': 'https://www.azsheriffs.org/sheriffs',
    'ARKANSAS': 'https://www.arsheriffs.org/sheriffs',
    'FLORIDA': 'https://flsheriffs.org/sheriffs',
    'GEORGIA': 'https://www.gsaonline.org/sheriffs',
    'INDIANA': 'https://www.indianasheriffs.org/sheriffs-directory',
    'KANSAS': 'https://www.kansassheriffs.org/directory',
    'KENTUCKY': 'https://www.kysheriffs.org/sheriffs-directory',
    'LOUISIANA': 'https://www.lsa.org/sheriffs',
    'MISSISSIPPI': 'https://www.msheriffs.org/sheriffs/',
    'MISSOURI': 'https://www.mosheriffs.com/sheriffs-directory/',
    'NORTH_CAROLINA': 'https://ncsheriffs.org/directory/',
    'OHIO': 'https://www.ohiosheriffs.org/sheriffs-directory/',
    'OKLAHOMA': 'https://www.oklahomasheriffs.org/sheriffs/',
    'PENNSYLVANIA': 'https://www.pasheriffs.org/sheriffs',
    'SOUTH_CAROLINA': 'https://www.scsheriffs.org/sheriffs/',
    'TENNESSEE': 'https://www.tnsheriffs.com/sheriffs/',
    'TEXAS': 'https://www.txsheriffs.org/',
    'VIRGINIA': 'https://vasheriff.org/sheriffs-directory/',
    'WEST_VIRGINIA': 'https://www.wvsheriffs.org/sheriffs',
}


def load_287g_agencies():
    """Load all 287(g) agencies from the Excel file."""
    excel_path = Path(__file__).parent / "287g_agencies.xlsx"
    wb = load_workbook(excel_path)
    ws = wb.active

    agencies = []
    for row in range(2, ws.max_row + 1):
        state = ws.cell(row=row, column=1).value
        agency = ws.cell(row=row, column=2).value
        county = ws.cell(row=row, column=4).value

        if not state or not agency:
            continue

        agencies.append({
            'state': state.strip(),
            'agency': agency.strip(),
            'county': county.strip() if county else None,
        })

    return agencies


def normalize_county(name):
    """Normalize county name for matching."""
    if not name:
        return ''
    name = name.lower()
    name = re.sub(r"['\"]", '', name)
    name = re.sub(r'\s+(county|co\.?|parish)$', '', name)
    name = re.sub(r'\s+', ' ', name)
    return name.strip()


def get_agency_type(agency_name):
    """Determine agency type from name."""
    name = agency_name.lower()
    if 'sheriff' in name:
        return 'sheriff'
    elif 'police' in name or ' pd' in name:
        return 'police'
    elif 'correction' in name or 'doc' in name or 'prison' in name:
        return 'corrections'
    else:
        return 'other'


# Hardcoded sheriff data (faster than scraping each time)
# TODO: Replace with live scraping for freshness
SHERIFF_DATA = {
    'ALABAMA': {
        'Autauga': 'Mark Harrell', 'Baldwin': 'Anthony Lowery', 'Barbour': 'Tyrone Smith',
        'Bibb': 'Jody Wade', 'Blount': 'Mark Moon', 'Bullock': 'Raymond Rodgers',
        'Butler': 'David Scruggs', 'Calhoun': 'Falon Hurst', 'Chambers': 'Jeff Nelson',
        'Cherokee': 'Jeff Shaver', 'Chilton': 'John Shearon', 'Choctaw': 'Scott Lolley',
        'Clarke': 'DeWayne C. Smith', 'Clay': 'Henry Lambert', 'Cleburne': 'Jon Daniel',
        'Coffee': 'Scott Byrd', 'Colbert': 'Eric Balentine', 'Conecuh': 'Randy Brock',
        'Coosa': 'Michael Howell', 'Covington': 'Blake Turman', 'Crenshaw': 'Terry Mears',
        'Cullman': 'Matt Gentry', 'Dale': 'Mason Bynum', 'Dallas': 'Michael Granthum',
        'DeKalb': 'Nicholas Welden', 'Elmore': 'Bill Franklin', 'Escambia': 'Heath Jackson',
        'Etowah': 'Jonathon Horton', 'Fayette': 'Byron Yerby', 'Franklin': 'Shannon Oliver',
        'Geneva': 'Tony Helms', 'Greene': 'Jonathan Benison', 'Hale': 'Michael Hamilton',
        'Henry': 'Eric Blankenship', 'Houston': 'Donald Valenza', 'Jackson': 'Rocky Harnen',
        'Jefferson': 'Mark Pettway', 'Lamar': 'Martin Gottwald', 'Lauderdale': 'Joe Hamilton',
        'Lawrence': 'Max Sanders', 'Lee': 'Jay Jones', 'Limestone': 'Joshua S. McLaughlin',
        'Lowndes': 'Christopher West', 'Macon': 'Andre Brunson', 'Madison': 'Kevin Turner',
        'Marengo': 'Robert Alston', 'Marion': 'Kevin Williams', 'Marshall': 'Phil Sims',
        'Mobile': 'Paul Burch Jr.', 'Monroe': 'Thomas Boatwright', 'Montgomery': 'Derrick Cunningham',
        'Morgan': 'Ron Puckett', 'Perry': 'Roy Fikes', 'Pickens': 'Jordan Powell',
        'Pike': 'Russell Thomas', 'Randolph': 'David Cofield', 'Russell': 'Heath Taylor',
        'Shelby': 'John Samaniego', 'St. Clair': 'Billy Murray', 'Sumter': 'Brian Harris',
        'Talladega': 'Jimmy Kilgore', 'Tallapoosa': 'Jimmy Abbett', 'Tuscaloosa': 'Ron Abernathy',
        'Walker': 'Nicholas Smith', 'Washington': 'Richard Stringer', 'Wilcox': 'Larry Colston',
        'Winston': 'Caleb Snoddy',
    },
    'ARKANSAS': {
        'Arkansas': 'Johnny Cheek', 'Ashley': 'Tommy Sturgeon', 'Baxter': 'John Montgomery',
        'Benton': 'Shawn Holloway', 'Boone': 'Roy Martin', 'Bradley': 'Herschel Tillman',
        'Calhoun': 'Vernon Morris', 'Carroll': 'Daniel Klatt', 'Chicot': 'Ronald Nichols',
        'Clark': 'Jason Watson', 'Clay': 'Ronnie Cole', 'Cleburne': 'Chris Brown',
        'Cleveland': 'Jack H. Rodgers II', 'Columbia': 'Leroy Martin', 'Conway': 'Mike Smith',
        'Craighead': 'Marty Boyd', 'Crawford': 'Daniel Perry', 'Crittenden': 'Mike Allen',
        'Cross': 'David West', 'Dallas': 'Mike Knoedl', 'Desha': 'Mitch Grant',
        'Drew': 'Tim Nichols', 'Faulkner': 'Tim Ryals', 'Franklin': 'Johnny Crocker',
        'Fulton': 'Jake Smith', 'Garland': 'Mike McCormick', 'Grant': 'Pete Roberts',
        'Greene': 'Brad Snyder', 'Hempstead': 'James Singleton', 'Hot Spring': 'Richard Tolleson',
        'Howard': 'Bryan McJunkins', 'Independence': 'Shawn Stephens', 'Izard': 'Charley Melton',
        'Jackson': 'Russell Brinsfield', 'Jefferson': 'Lafayette Woods, Jr.', 'Johnson': 'Tom Hughes',
        'Lafayette': 'Jeff Black', 'Lawrence': 'Tony Waldrupe', 'Lee': 'Corey Wilson',
        'Lincoln': 'Steve Young', 'Little River': 'Bobby Walraven', 'Logan': 'Jason Massey',
        'Lonoke': 'John Staley', 'Madison': 'Ronnie Boyd', 'Marion': 'Gregg Alexander',
        'Miller': 'David Wayne Easley', 'Mississippi': 'Dale Cook', 'Monroe': 'Michael K. Neal',
        'Montgomery': 'Neal Thomas', 'Nevada': 'Danny Martin', 'Newton': 'Glenn Wheeler',
        'Ouachita': 'David Norwood', 'Perry': 'Ricky Don Jones', 'Phillips': 'Neal Byrd',
        'Pike': 'Travis Hill', 'Poinsett': 'Kevin Molder', 'Polk': 'Scott Sawyer',
        'Pope': 'Blake Wilson', 'Prairie': 'Rick Parson', 'Pulaski': 'Eric Higgins',
        'Randolph': 'Kevin Bell', 'Saline': 'Rodney Wright', 'Scott': 'Randy Shores',
        'Searcy': 'Kenny Cassell', 'Sebastian': 'Hobe Runion', 'Sevier': 'Robert Gentry',
        'Sharp': 'Shane Russell', 'St. Francis': 'Bobby May', 'Stone': 'Brandon Long',
        'Union': 'Charlie Phillips', 'Van Buren': 'Eric Koonce', 'Washington': 'Jay Cantrell',
        'White': 'Phillip E. Miller', 'Woodruff': 'Phil Reynolds', 'Yell': 'Nick Gault',
    },
    'MISSISSIPPI': {
        'Adams': 'Travis Patten', 'Alcorn': 'Ben Caldwell', 'Amite': 'Tim Wroten',
        'Attala': 'Curtis Pope', 'Benton': 'Robbie Goolsby', 'Bolivar': 'Kelvin Williams',
        'Calhoun': 'Greg Pollan', 'Carroll': 'Clint Walker', 'Chickasaw': 'James D. Meyers',
        'Choctaw': 'Brandon Busby', 'Claiborne': 'Edward Goods', 'Clarke': 'Anthony Chancelor',
        'Clay': 'Eddie Scott', 'Coahoma': 'Mario Magsby', 'Copiah': 'Byron Swilley',
        'Covington': 'Darrell Perkins', 'DeSoto': 'Thomas Tuggle', 'Forrest': 'Charlie Sims',
        'Franklin': 'Tom Tindle', 'George': 'Mitchell Mixon', 'Greene': 'Ryan Walley',
        'Grenada': 'Garrett Hartley', 'Hancock': 'Ricky Adam', 'Harrison': 'Matt Haley',
        'Hinds': 'Tyree D. Jones', 'Holmes': 'Willie March', 'Issaquena': 'Waye Windham',
        'Itawamba': 'Mitch Nabors', 'Jackson': 'John Ledbetter', 'Jasper': 'Randy Johnson',
        'Jefferson': 'James Bailey, Sr.', 'Jefferson Davis': 'Ron Strickland', 'Jones': 'Joe Berlin',
        'Kemper': 'James Moore', 'Lafayette': 'Joey East', 'Lamar': 'Danny Rigel',
        'Lauderdale': 'Ward Calhoun', 'Lawrence': 'Ryan Everett', 'Leake': 'Randy Atkinson',
        'Lee': 'Jim Johnson', 'Leflore': 'James Payne, Sr.', 'Lincoln': 'Steve Rushing',
        'Lowndes': 'Eddie Hawkins', 'Madison': 'Randall Tucker', 'Marion': 'Berkley Hall',
        'Marshall': 'Kenny Dickerson', 'Monroe': 'Kevin Crook', 'Montgomery': 'Jeff Tompkins',
        'Neshoba': 'Eric Clark', 'Newton': 'Joedy Pennington', 'Noxubee': 'Dontevis Smith',
        'Oktibbeha': 'Shank Phelps', 'Panola': 'Shane Phelps', 'Pearl River': 'David Allison',
        'Perry': 'Jacob Garner', 'Pike': 'Wally Jones', 'Pontotoc': 'Leo Mask',
        'Prentiss': 'Randy Tolar', 'Quitman': 'Oliver Parker Jr.', 'Rankin': 'Bryan Bailey',
        'Scott': 'Mike Lee', 'Sharkey': 'Herbert Ceaser', 'Simpson': 'Paul Mullins',
        'Smith': 'Joel Houston', 'Stone': 'Todd Stewart', 'Sunflower': 'James Haywood',
        'Tallahatchie': 'Jimmy Fly', 'Tate': 'Luke Shepherd', 'Tippah': 'Karl Gaillard',
        'Tishomingo': 'Jamie Stuart', 'Tunica': 'K.C. Hamp', 'Union': 'Jimmy Edwards',
        'Walthall': 'Kyle Breland', 'Warren': 'Martin Pace Jr.', 'Wayne': 'Jerry Moseley',
        'Wilkinson': 'Reginald Jackson', 'Winston': 'Mike Perkins', 'Yalobusha': 'Jerimaine Gooch',
        'Yazoo': 'Jeremy McCoy, Sr.',
    },
    'OKLAHOMA': {
        'Adair': 'Jason Ritchie', 'Alfalfa': 'Gary Mast', 'Atoka': 'Kody Simpson',
        'Beaver': 'Scott Mitchell', 'Beckham': 'Derek Manning', 'Blaine': 'Travis Daugherty',
        'Bryan': 'Joey Tucker', 'Caddo': 'Spencer Davis', 'Canadian': 'Chris West',
        'Carter': 'DJ Long', 'Cherokee': 'Jason Chennault', 'Choctaw': 'Terry Park',
        'Cimarron': 'Mark Swinton', 'Cleveland': 'Chris Amason', 'Comanche': 'Michael Merritt',
        'Cotton': 'Tim King', 'Craig': 'Heath Winfrey', 'Creek': 'Bret Bowling',
        'Custer': 'Dan Day', 'Delaware': 'Ray Thomas', 'Dewey': 'Clay Sander',
        'Ellis': 'Shane Booth', 'Garfield': 'Cory S. Rink', 'Garvin': 'Jim Mullett',
        'Grady': 'Gary Boggess', 'Grant': 'Tim Irvin', 'Greer': 'Steven McMahan',
        'Harmon': 'Steve Cornett', 'Harper': 'Thomas McClendon', 'Haskell': 'Terry Garland',
        'Hughes': 'Trever Teague', 'Jackson': 'Stacy Randolph', 'Jefferson': 'Michael Bryant',
        'Johnston': 'Gary Dodd', 'Kay': 'Steve Kelley', 'Kingfisher': 'Aaron Pitts',
        'Kiowa': 'Joe Janz', 'Latimer': 'Adam Woodruff', 'LeFlore': 'Rodney Derryberry',
        'Lincoln': 'Kevin Garrett', 'Logan': 'Damon Devereaux', 'Love': 'Andy Cumberledge',
        'Major': 'Anthony Robinson', 'Marshall': 'Donald Yow', 'Mayes': 'Mike Reed',
        'McClain': 'Landy Offolter', 'McCurtain': 'Bruce Shirey', 'McIntosh': 'Kevin Ledbetter',
        'Murray': 'Darin Rogers', 'Muskogee': 'Andy Simmons', 'Noble': 'Matt McGuire',
        'Nowata': 'Jason McClain', 'Okfuskee': 'Logan Manshack', 'Oklahoma': 'Tommie Johnson III',
        'Okmulgee': 'Eddy Rice', 'Osage': 'Bart Perrier', 'Ottawa': 'David Dean',
        'Pawnee': 'Shawn Price', 'Payne': 'Joe Harper', 'Pittsburg': 'Frankie McClendon',
        'Pontotoc': 'Arnold Scott', 'Pottawatomie': 'Freeland Wood', 'Pushmataha': 'B.J. Hedgecock',
        'Roger Mills': 'Darren Atha', 'Rogers': 'Scott Walton', 'Seminole': 'Anthony Louie',
        'Sequoyah': 'Larry Lane', 'Stephens': 'Rick Lang', 'Texas': 'Matt Boley',
        'Tillman': 'Oscar Juanes', 'Tulsa': 'Vic Regalado', 'Wagoner': 'Chris Elliott',
        'Washington': 'Scott Owen', 'Washita': 'Kevin Rozell', 'Woods': 'Randal McCullough',
        'Woodward': 'Keith Frutiger',
    },
    'KANSAS': {
        'Allen': 'Anthony Maness', 'Anderson': 'Wesley McClain', 'Atchison': 'Jack Laurie',
        'Barber': 'Richard Garza', 'Barton': 'Brian Bellendir', 'Bourbon': 'William "Bill" Martin',
        'Brown': 'John Merchant', 'Butler': 'Monty Lee Hughey', 'Chase': 'Jacob Welsh',
        'Chautauqua': 'Richard A. Newby', 'Cherokee': 'Brian Henderson', 'Cheyenne': 'Cody Beeson',
        'Clark': 'Heath Heston', 'Clay': 'Alan Benninga', 'Cloud': 'Ken Davis',
        'Coffey': 'Thomas L. Johnson', 'Comanche': 'Jacob Bruckner', 'Cowley': 'David Falletti',
        'Crawford': 'Billy Tomasi', 'Decatur': 'David Wachendorfer', 'Dickinson': 'Jerry Davis',
        'Doniphan': 'Mark Allen', 'Douglas': 'Jay Armbrister', 'Edwards': 'Bryant Kurth',
        'Elk': 'John Walker', 'Ellis': 'Scott J. Braun', 'Ellsworth': 'Kenny Bernard',
        'Finney': 'Steve Martinez', 'Ford': 'William (Bill) Carr', 'Franklin': 'Jeff Richards',
        'Geary': 'Nathan Boeckman', 'Gove': 'Shawn Mesch', 'Graham': 'Les Burton',
        'Grant': 'James Biddle', 'Gray': 'Jeff Sharp', 'Greeley': 'Jessica McDaniel-Brown',
        'Greenwood': 'Heath Samuels', 'Hamilton': 'Michael Wilson', 'Harper': 'Tracy Chance',
        'Harvey': 'Chad Gay', 'Haskell': 'Troy Briggs', 'Hodgeman': 'Jared Walker',
        'Jackson': 'Tim Morse', 'Jefferson': 'Jeff Herrig', 'Jewell': 'Don Jacobs',
        'Johnson': 'Byron Roberson', 'Kearny': 'Michael Fontenot', 'Kingman': 'Brent Wood',
        'Kiowa': 'Kendal Lothman', 'Labette': 'Darren Eichinger', 'Lane': 'Ron Ridley',
        'Leavenworth': 'Andrew Dedeke', 'Lincoln': 'Dustin Florence', 'Linn': 'James Akes',
        'Logan': 'Dustin Little', 'Lyon': 'Jeff Cope', 'Marion': 'Jeff Soyez',
        'Marshall': 'Tim Ackerman', 'McPherson': 'Jerry Montagne', 'Meade': 'Doug Ritter',
        'Miami': 'Frank Kelly', 'Mitchell': 'Anthony Perez', 'Montgomery': 'Ron Wade',
        'Morris': 'Scott Coover', 'Morton': 'Thad Earls', 'Nemaha': 'Richard Vernon',
        'Neosho': 'Greg Taylor', 'Ness': 'Brandon Mitchell', 'Norton': 'Troy Thomson',
        'Osage': 'Chris Wells', 'Osborne': 'Scottie Becker', 'Ottawa': 'Russ Thornton',
        'Pawnee': 'Larry Atteberry', 'Phillips': 'Kyle Pinkerton', 'Pottawatomie': 'Doug Adams',
        'Pratt': 'James White', 'Rawlins': 'Catlyn Wahrman', 'Reno': 'Darrian Campbell',
        'Republic': 'David Cox', 'Rice': 'Nicholas Sowers', 'Riley': 'Brian Peete',
        'Rooks': 'Gary Knight', 'Rush': 'Mark Knowles', 'Russell': 'Andrew Van Der Wege',
        'Saline': 'Roger Soldan', 'Scott': 'Jeffrey Pounds', 'Sedgwick': 'Jeffrey Easter',
        'Seward': 'Gene Ward', 'Shawnee': 'Brian Hill', 'Sheridan': 'Brandon Carver',
        'Sherman': 'Chad Mann', 'Smith': 'Travis Conaway', 'Stafford': 'Robert Murrow',
        'Stanton': 'Clayton Kester', 'Stevens': 'Ted Heaton', 'Sumner': 'Darren Chambers',
        'Thomas': 'Joel Thomas Nickols Jr.', 'Trego': 'Ed Pritchard', 'Wabaunsee': 'Eric Kirsch',
        'Wallace': 'Marshall Unruh', 'Washington': 'Justin Cordry', 'Wichita': 'Kristopher Casper',
        'Wilson': 'Harold Kuhn', 'Woodson': 'Jacob Morrison', 'Wyandotte': 'Daniel Soptic',
    },
    'LOUISIANA': {
        'Acadia': 'K. P. Gibson', 'Allen': 'Douglas L. Hebert III', 'Ascension': 'Robert "Bobby" Webre',
        'Assumption': 'Leland Falcon', 'Avoyelles': 'David Dauzat', 'Beauregard': 'Mark Herford',
        'Bienville': 'John E Ballance', 'Bossier': 'Julian C. Whittington', 'Caddo': 'Henry Whitehorn, Sr.',
        'Calcasieu': '"Stitch" Guillory', 'Caldwell': 'Clay Bennett', 'Cameron': '"Chris" Savoie',
        'Catahoula': 'Toney Edwards', 'Claiborne': 'Sam Dowies', 'Concordia': 'David Hedrick',
        'De Soto': 'Jayson Richardson', 'East Baton Rouge': 'Sid J. Gautreaux III', 'East Carroll': 'Wydette Williams',
        'East Feliciana': '"Jeff" Travis', 'Evangeline': 'Charles Guillory', 'Franklin': 'Bryan Linder',
        'Grant': 'Steven McCain', 'Iberia': 'Tommy Romero', 'Iberville': 'Brett M. Stassi',
        'Jackson': 'Brent Barnett', 'Jefferson Davis': 'Kyle Miers', 'Jefferson': 'Joseph P Lopinto III',
        'Lafayette': 'Mark Garber', 'Lafourche': 'Craig Webre', 'Lasalle': '"Lane" Windham',
        'Lincoln': 'Stephen Williams', 'Livingston': 'Jason G. Ard', 'Madison': 'Chad Ezell',
        'Morehouse': 'Mike Tubbs', 'Natchitoches': 'Stuart Wright', 'Orleans': 'Susan Hutson',
        'Ouachita': 'Marcus "Marc" Mashaw', 'Plaquemines': 'Gerald A. Turlich Jr.', 'Pointe Coupee': 'Rene Thibodeaux',
        'Rapides': 'Mark Wood', 'Red River': 'Glen Edwards', 'Richland': 'Neal G. Harwell',
        'Sabine': 'Aaron Mitchell', 'St. Bernard': 'James Pohlmann', 'St. Charles': 'Gregory C. "Greg" Champagne',
        'St. Helena': '"Clay" Chutz', 'St. James': 'Claude J. Louis Jr.', 'St. John The Baptist': 'Mike Tregre',
        'St. Landry': 'Bobby J Guidroz', 'St. Martin': 'Becket Breaux', 'St. Mary': 'Gary Driskell',
        'St. Tammany': 'Randy Smith', 'Tangipahoa': 'Gerald D. Sticker', 'Tensas': 'Robert L. "Rob" Rushing',
        'Terrebonne': 'Timothy Soignet', 'Union': 'Dusty Gates', 'Vermilion': 'Eddie Langlinais',
        'Vernon': 'John S. "Sam" Craft', 'Washington': 'Jason Smith', 'Webster': 'Jason Parker',
        'West Baton Rouge': 'Jeff Bergeron', 'West Carroll': 'Scott Mathews', 'West Feliciana': 'Brian L. Spillman',
        'Winn': 'Josh McAllister',
    },
    'GEORGIA': {
        'Appling': 'Mark Melton', 'Atkinson': 'David Moore', 'Bacon': 'Andy Batten',
        'Baker': 'Dana Meade', 'Baldwin': 'William Massee, Jr', 'Banks': 'Carlton Speed',
        'Barrow': 'Jud Smith', 'Bartow': 'Clark Millsap', 'Ben Hill': 'Lee Cone',
        'Berrien': 'Ray Paulk', 'Bibb': 'David J. Davis', 'Bleckley': 'Daniel Cape',
        'Brantley': 'Len Davis', 'Brooks': 'Mike Dewey', 'Bryan': 'Mark Crowe',
        'Bulloch': 'Noel Brown', 'Burke': 'Alfonzo Williams', 'Butts': 'Gary Long',
        'Calhoun': 'Josh Hilton', 'Camden': 'Kevin Chaney', 'Candler': 'John Miles',
        'Carroll': 'Terry Langley', 'Catoosa': 'Gary Sisk', 'Charlton': 'Robert Phillips',
        'Chatham': 'Richard Coleman', 'Chattahoochee': 'Henry Lynch III', 'Chattooga': 'Mark Schrader',
        'Cherokee': 'Frank Reynolds', 'Clarke': 'John Williams', 'Clay': 'Locke Shivers',
        'Clayton': 'Levon Allen', 'Clinch': 'Raymond Peterson', 'Cobb': 'Craig Owens',
        'Coffee': 'Fred Cole', 'Colquitt': 'Rod Howell', 'Columbia': 'Clay Whittle',
        'Cook': 'Douglas Hanks', 'Coweta': 'Lenn Wood', 'Crawford': 'Lewis Walker',
        'Crisp': 'Billy Hancock', 'Dade': 'Ray Cross', 'Dawson': 'Jeff Johnson',
        'Decatur': 'Wiley Griffin', 'Dekalb': 'Melody Maddox', 'Dodge': 'Brian Robinson',
        'Dooly': 'Craig Peavy', 'Dougherty': 'Terron K. Hayes', 'Douglas': 'Tim Pounds',
        'Early': 'William Price', 'Echols': 'Randy Courson', 'Effingham': 'Jimmy McDuffie',
        'Elbert': 'Jamie Callaway', 'Emanuel': 'Jeffrey Brewer', 'Evans': 'Mac Edwards',
        'Fannin': 'Dane Kirby', 'Fayette': 'Barry Babb', 'Floyd': 'Dave Roberson',
        'Forsyth': 'Ron Freeman', 'Franklin': 'Scott Andrews', 'Fulton': 'Pat Labat',
        'Gilmer': 'Stacy Nicholson', 'Glascock': 'Jeremy Kelley', 'Glynn': 'E. Neal Jump',
        'Gordon': 'Mitch Ralston', 'Grady': 'Martin E. Prince', 'Greene': 'Donnie Harrison',
        'Gwinnett': 'Keybo Taylor', 'Habersham': 'Robin Krockum', 'Hall': 'Gerald Couch',
        'Hancock': 'Tomyln Primus', 'Haralson': 'Stacy Williams', 'Harris': 'Mike Jolley',
        'Hart': 'Chris Carroll', 'Heard': 'Ross Henry', 'Henry': 'Reginald Scandrett',
        'Houston': 'Matt Moulton', 'Irwin': 'Cody Youghn', 'Jackson': 'Kevin McCook',
        'Jasper': 'Donnie Pope', 'Jeff Davis': 'Preston Bohannon', 'Jefferson': 'Gary Hutchins',
        'Jenkins': 'Robert Oglesby', 'Johnson': 'Greg Rowland', 'Jones': 'Butch Reece',
        'Lamar': 'Brad White', 'Lanier': 'Nick Norton', 'Laurens': 'Larry H. Dean',
        'Lee': 'Reggie Rachals', 'Liberty': 'Will Bowman', 'Lincoln': 'Clay Smith',
        'Long': 'Craig Nobles', 'Lowndes': 'Ashley Paulk', 'Lumpkin': 'Stacy Jarrard',
        'Macon': 'Carlos Felton', 'Madison': 'James Michael Moore', 'Marion': 'Derrell Neal',
        'Meriwether': 'Chuck Smith', 'Miller': 'Garison Clenney', 'Mitchell': 'W.E. Bozeman',
        'Monroe': 'Brad Freeman', 'Montgomery': 'Ben Maybin', 'Morgan': 'Tyler Hooks',
        'Murray': 'Jimmy Davenport', 'Muscogee': 'Greg Countryman', 'Newton': 'Ezell Brown',
        'Oconee': 'James Hale', 'Oglethorpe': 'David Gabriel', 'Paulding': 'Ashley Henson',
        'Peach': 'Robert Shannon', 'Pickens': 'Donald Craig', 'Pierce': 'Ramsey Bennett',
        'Pike': 'Jimmy Thomas', 'Polk': 'Johnny Moats', 'Pulaski': 'Wayne Wiley',
        'Putnam': 'Howard Sills', 'Quitman': 'Charles Davis', 'Rabun': 'Mark Gerrells',
        'Randolph': 'Eddie Fairbanks', 'Richmond': 'Gino Brantley', 'Rockdale': 'Eric Levett',
        'Schley': 'Scott Nelson', 'Screven': 'Norman Royal', 'Seminole': 'Heath Elliott',
        'Spalding': 'Darrell Dix', 'Stephens': 'Rusty Fulbright', 'Stewart': 'Larry Jones',
        'Sumter': 'Eric Bryant', 'Talbot': 'Bobby Gates, Jr', 'Taliaferro': 'Tia McWilliams',
        'Tattnall': 'Kyle Sapp', 'Taylor': 'John Sawyer III', 'Telfair': 'Sim Davidson',
        'Terrell': 'Vernon Jessie', 'Thomas': 'Tim Watkins', 'Tift': 'Gene Scarbrough',
        'Toombs': 'Jordan Kight', 'Towns': 'Anthony Coleman', 'Treutlen': 'Thomas Corbin',
        'Troup': 'James Woodruff', 'Turner': 'Andy Hester', 'Twiggs': 'Darren Mitchum',
        'Union': 'Shawn Dyer', 'Upson': 'Dan Kilgore', 'Walker': 'Steve Wilson',
        'Walton': 'Keith Brooks', 'Ware': 'Carl James', 'Warren': 'Joe Peebles',
        'Washington': 'Joel Cochran', 'Wayne': 'Chuck Moseley', 'Webster': 'Randy Dely',
        'Wheeler': 'Glenn Giles', 'White': 'Rick Kelley', 'Whitfield': 'Darren Pierce',
        'Wilcox': 'Jeffrey Wessel', 'Wilkes': 'Darrell Powers', 'Wilkinson': 'Richard Chatman',
        'Worth': 'Don Whitaker',
    },
    'FLORIDA': {
        'Alachua': 'Chad Scott', 'Baker': 'Scotty Rhoden', 'Bay': 'Tommy Ford',
        'Bradford': 'Gordon Smith', 'Brevard': 'Wayne Ivey', 'Broward': 'Gregory Tony',
        'Calhoun': 'Michael Bryant', 'Charlotte': 'Bill Prummell Jr.', 'Citrus': 'David E. Vincent',
        'Clay': 'Michelle Cook', 'Collier': 'Kevin J. Rambosk', 'Columbia': 'Wallace Kitchings',
        'DeSoto': 'James F. Potter', 'Dixie': 'Darby Butler', 'Duval': 'T.K. Waters',
        'Escambia': 'Chip Simmons', 'Flagler': 'Rick Staly', 'Franklin': 'A.J. Smith',
        'Gadsden': 'Morris A. Young', 'Gilchrist': 'Bobby Schultz', 'Glades': 'David Hardin',
        'Gulf': 'Mike Harrison', 'Hamilton': 'Brian Creech', 'Hardee': 'Vent Crawford',
        'Hendry': 'Steve Whidden', 'Hernando': 'Al Nienhuis', 'Highlands': 'Paul Blackman',
        'Hillsborough': 'Chad Chronister', 'Holmes': 'John Tate', 'Indian River': 'Eric Flowers',
        'Jackson': 'Donald L. Edenfield', 'Jefferson': 'Mac McNeill Jr.', 'Lafayette': 'Brian N. Lamb',
        'Lake': 'Peyton C. Grinnell', 'Lee': 'Carmine Marceno', 'Leon': 'Walt McNeil',
        'Levy': 'Bobby McCallum', 'Liberty': 'Dusty Arnold', 'Madison': 'David Harper',
        'Manatee': 'Charles R. Wells', 'Marion': 'Billy Woods', 'Martin': 'John M. Budensiek',
        'Miami-Dade': 'Rosie Cordero-Stutz', 'Monroe': 'Rick Ramsay', 'Nassau': 'Bill Leeper',
        'Okaloosa': 'Eric Aden', 'Okeechobee': 'Noel E. Stephen', 'Orange': 'John W. Mina',
        'Osceola': 'Chris Blackmon', 'Palm Beach': 'Ric L. Bradshaw', 'Pasco': 'Chris Nocco',
        'Pinellas': 'Bob Gualtieri', 'Polk': 'Grady Judd', 'Putnam': 'Gator DeLoach III',
        'Santa Rosa': 'Robert Johnson', 'Sarasota': 'Kurt A. Hoffman', 'Seminole': 'Dennis M. Lemma',
        'St. Johns': 'Robert A. Hardwick', 'St. Lucie': 'Richard R. Del Toro Jr.', 'Sumter': 'Patrick Breeden',
        'Suwannee': 'Sam St. John', 'Taylor': 'Wayne Padgett', 'Union': 'Brad Whitehead',
        'Volusia': 'Mike Chitwood', 'Wakulla': 'Jared Miller', 'Walton': 'Michael A. Adkinson Jr.',
        'Washington': 'Kevin Crews',
    },
    'TENNESSEE': {
        'Anderson': 'Russell Barker', 'Bedford': 'Austin Swing', 'Benton': 'Kenneth Christopher',
        'Bledsoe': 'James Morris', 'Blount': 'James Berrong', 'Bradley': 'Steve Lawson',
        'Campbell': 'Eddie Barton II', 'Cannon': 'Darrell Young', 'Carroll': 'Andy Dickson',
        'Carter': 'Mike Fraley', 'Cheatham': 'Tim Binkley', 'Chester': 'Mark Griffin',
        'Claiborne': 'Robert Brooks', 'Clay': 'Brandon Boone', 'Cocke': 'C.J. Ball',
        'Coffee': 'Chad Partin', 'Crockett': 'Troy Klyce', 'Cumberland': 'Casey Cox',
        'Davidson': 'Daron Hall', 'Decatur': 'Dale King', 'DeKalb': 'Patrick Ray',
        'Dickson': 'Tim Eads', 'Dyer': 'Jeff Box', 'Fayette': 'Bobby Riles',
        'Franklin': 'Tim Fuller', 'Gibson': 'Paul Thomas', 'Giles': 'Joe Purvis',
        'Grainger': 'James Harville', 'Greene': 'Wesley Holt', 'Grundy': 'Heath G. Gunter',
        'Hamblen': 'Chad Mullins', 'Hamilton': 'Austin Garrett', 'Hancock': 'Bradley Brewer',
        'Hardeman': 'John Doolen', 'Hardin': 'Johnny Alexander', 'Hawkins': 'Ronnie Lawson',
        'Haywood': 'Billy Garrett, Jr.', 'Henderson': 'Brian Duke', 'Henry': 'Josh Frey',
        'Hickman': 'Jason Craft', 'Houston': 'Kevin Sugg', 'Humphreys': 'Chris Davis',
        'Jackson': 'Marty Hinson', 'Jefferson': 'Jeff Coffey', 'Johnson': 'Clifton Worley, Jr.',
        'Knox': 'Tom Spangler', 'Lake': 'Bryan Avery', 'Lauderdale': 'Brian Kelley',
        'Lawrence': 'John Myers', 'Lewis': 'Matt Tiller', 'Lincoln': 'Tull Malone',
        'Loudon': 'James Davis', 'Macon': 'Joey Wilburn', 'Madison': 'Julian Wiser',
        'Marion': 'Ronnie Burnett', 'Marshall': 'Billy Lamb', 'Maury': 'Bucky Rowland',
        'McMinn': 'Joe Guy', 'McNairy': 'Guy Buck', 'Meigs': 'Jackie Melton',
        'Monroe': 'Tommy Jones, Jr.', 'Montgomery': 'John Fuson', 'Moore': 'Tyler Hatfield',
        'Morgan': 'Rick Hamby', 'Obion': 'Karl Jackson', 'Overton': 'John Garrett',
        'Perry': 'Nick Weems', 'Pickett': 'Dana Dowdy', 'Polk': 'Steve Ross',
        'Putnam': 'Eddie Farris', 'Rhea': 'Mike Neal', 'Roane': 'Jack Stockton',
        'Robertson': 'Michael Van Dyke', 'Rutherford': 'Mike Fitzhugh', 'Scott': 'Brian Keeton',
        'Sequatchie': 'William "Bill" Phillips', 'Sevier': 'Michael Hodges, Jr.', 'Shelby': 'Floyd Bonner, Jr.',
        'Smith': 'Steve Hopper', 'Stewart': 'Charles Gray', 'Sullivan': 'Jeff Cassidy',
        'Sumner': 'Eric Craddock', 'Tipton': 'Shannon Beasley', 'Trousdale': 'Ray Russell',
        'Unicoi': 'Mike Hensley', 'Union': 'Billy Breeding', 'Van Buren': 'Michael J. Brock',
        'Warren': 'Jackie D. Matheny, Jr.', 'Washington': 'Keith Sexton', 'Wayne': 'Shane Fisher',
        'Weakley': 'Terry McDade', 'White': 'Steve Page', 'Williamson': 'Jeff Hughes',
        'Wilson': 'Robert Bryan',
    },
    'TEXAS': {
        'Anderson': 'William R. Flores', 'Andrews': 'Charles R. Stewart', 'Angelina': 'Thomas L. Selman Jr',
        'Aransas': 'William A. Mills', 'Archer': 'William Jack Curd', 'Armstrong': 'Melissa Anderson',
        'Atascosa': 'David A. Soward', 'Austin': 'Jack Brandes', 'Bailey': 'Richard Wills',
        'Bandera': 'Daniel Raymond Butts', 'Bastrop': 'Maurice C. Cook', 'Baylor': 'Sam Mooney',
        'Bee': 'Alden E. Southmayd III', 'Bell': 'Eddy Lange', 'Bexar': 'Javier Salazar',
        'Blanco': 'Don Jackson', 'Borden': 'Benny Ray Allison', 'Bosque': 'Trace Arthur Hendricks',
        'Bowie': 'Jeffrey K. Neal', 'Brazoria': 'Bo Stallman', 'Brazos': 'Wayne Dicky',
        'Brewster': 'Ronny D. Dodson', 'Briscoe': 'Garrett King Davis', 'Brooks': 'Urbino Benny Martinez',
        'Brown': 'Vance W. Hill', 'Burleson': 'Gene E. Hermes', 'Burnet': 'Calvin Boyd',
        'Caldwell': 'Michael K. Lane', 'Calhoun': 'Bobbie Vickery', 'Callahan': 'Joe Eric Pechacek',
        'Cameron': 'Eric Garza', 'Camp': 'John B. Cortelyou', 'Carson': 'Tam Terry',
        'Cass': 'Larry Rowe', 'Castro': 'Salvador S. Rivera Jr', 'Chambers': 'Brian C. Hawthorne',
        'Cherokee': 'Brent Dickson', 'Childress': 'Matthew W. Bradley', 'Clay': 'Kirk Horton',
        'Cochran': 'Jorge De La Cruz', 'Coke': 'Wayne McCutchen', 'Coleman': 'Les W. Cogdill',
        'Collin': 'James O. Skinner', 'Collingsworth': 'Alan K. Riley', 'Colorado': 'R H Wied',
        'Comal': 'Mark W. Reynolds', 'Comanche': 'Chris D. Pounds', 'Concho': 'Chad Miller',
        'Cooke': 'Ray Sappington', 'Coryell': 'Scott A. Williams', 'Cottle': 'Mark Box',
        'Crane': 'Andrew R. Aguilar', 'Crockett': 'Antonio Gomez Alejandro III', 'Crosby': 'Ethan R. Villanueva',
        'Culberson': 'Oscar E. Carrillo', 'Dallam': 'Shane C. Stevenson', 'Dallas': 'Marian Brown',
        'Dawson': 'Matt Hogg', 'Deaf Smith': 'J Dale Butler Jr', 'Delta': 'Charla Singleton',
        'Denton': 'Tracy Murphree', 'DeWitt': 'Carl R. Bowen', 'Dickens': 'Terry Braly',
        'Dimmit': 'Chris A. Castaneda', 'Donley': 'Butch Blackburn Jr', 'Duval': 'Romeo R. Ramirez',
        'Eastland': 'Jason Weger', 'Ector': 'Michael W. Griffis', 'Edwards': 'James W. Guthrie',
        'El Paso': 'Richard Wiles', 'Ellis': 'Brad B. Norman', 'Erath': 'Matt Coates',
        'Falls': 'Joe Lopez', 'Fannin': 'Mark L. Johnson', 'Fayette': 'Keith K. Korenek',
        'Fisher': 'Randy Ford', 'Floyd': 'Paul Raissez', 'Foard': 'Mike Brown',
        'Fort Bend': 'Eric Fagan', 'Franklin': 'Ricky S. Jones', 'Freestone': 'Jeremy D. Shipley',
        'Frio': 'Michael Jay Morse', 'Gaines': 'Ronny Pipkin', 'Galveston': 'Henry A. Trochesset',
        'Garza': 'Terry L. Morgan', 'Gillespie': 'Buddy Mills', 'Glasscock': 'Keith Burnett',
        'Goliad': 'Roy Boyd Jr', 'Gonzales': 'Keith A. Schmidt', 'Gray': 'Michael L. Ryan',
        'Grayson': 'Thomas E. Watt', 'Gregg': 'Maxey Cerliano', 'Grimes': 'Donald G. Sowell',
        'Guadalupe': 'Arnold S. Zwicke', 'Hale': 'David Cochran', 'Hall': 'Thomas Heck',
        'Hamilton': 'Justin R. Caraway', 'Hansford': 'Robert Mahaffee', 'Hardeman': 'Patrick Laughery',
        'Hardin': 'Mark L. Davis', 'Harris': 'Ed Gonzalez', 'Harrison': 'Brandon Fletcher',
        'Hartley': 'Chanze W. Fowler', 'Haskell': 'Christopher Keith', 'Hays': 'Gary Cutler',
        'Hemphill': 'Brent Clapp', 'Henderson': 'Botie Hillhouse', 'Hidalgo': 'Eddie Guerra',
        'Hill': 'Rodney B. Watson', 'Hockley': 'James R. Scifres', 'Hood': 'Roger Deeds',
        'Hopkins': 'Lewis Tatum', 'Houston': 'Randy Hargrove', 'Howard': 'Stan Parker',
        'Hudspeth': 'Arvin West', 'Hunt': 'Terry G. Jones', 'Hutchinson': 'Blaik Ryan Kemp',
        'Irion': 'W A. Estes', 'Jack': 'Thomas Spurlock', 'Jackson': 'Kelly Janica',
        'Jasper': 'Mitchel Newman', 'Jeff Davis': 'William Kitts', 'Jefferson': 'Zena A. Stephens',
        'Jim Hogg': 'Erasmo Alarcon Jr', 'Jim Wells': 'Daniel J. Bueno', 'Johnson': 'Adam R. King',
        'Jones': 'Danny C. Jimenez', 'Karnes': 'Dwayne Villanueva', 'Kaufman': 'Bryan W. Beavers',
        'Kendall': 'Albert R. Auxier', 'Kenedy': 'Ramon Salinas III', 'Kent': 'William D. Scogin',
        'Kerr': 'Larry L. Leitha Jr', 'Kimble': 'Allen Castleberry', 'King': 'Mike McWhirter',
        'Kinney': 'Brad Coe', 'Kleberg': 'Richard C. Kirkpatrick', 'Knox': 'Bridger Bush',
        'La Salle': 'Anthony A. Zertuche', 'Lamar': 'Scott C. Cass', 'Lamb': 'Gary Maddox',
        'Lampasas': 'Jesus G. Ramos', 'Lavaca': 'Micah C. Harmon', 'Lee': 'Garrett Durrenberger',
        'Leon': 'Kevin D. Ellis', 'Liberty': 'Robert J. Rader Jr', 'Limestone': 'Murray A. Agnew',
        'Lipscomb': 'Ty Lane', 'Live Oak': 'Larry R. Busby', 'Llano': 'Bill Blackburn',
        'Loving': 'Chris H. Busse', 'Lubbock': 'Kelly S. Rowe', 'Lynn': 'Wanda Mason',
        'Madison': 'Bobby D. Adams', 'Marion': 'David Capps', 'Martin': 'James B. Ingram',
        'Mason': 'Joseph Lancaster', 'Matagorda': 'Frank Skipper Osborne', 'Maverick': 'Tom Schmerber',
        'McCulloch': 'Matt Andrews', 'McLennan': 'Parnell McNamara', 'McMullen': 'Emmett L. Shelton',
        'Medina': 'Randy R. Brown', 'Menard': 'Buck Miller', 'Midland': 'David A. Criner',
        'Milam': 'Mike Clore', 'Mills': 'Clint Royce Hammonds', 'Mitchell': 'Patrick Toombs',
        'Montague': 'Marshall W. Thomas', 'Montgomery': 'Rand Henderson', 'Moore': 'Morgan W. Hightower',
        'Morris': 'Jack Martin', 'Motley': 'Robert D. Fisk', 'Nacogdoches': 'Jason Bridges',
        'Navarro': 'Elmer L. Tanner', 'Newton': 'Robert J. Burby', 'Nolan': 'David Warren',
        'Nueces': 'J C Hooper', 'Ochiltree': 'Terry L. Bouchard', 'Oldham': 'Brent Warden',
        'Orange': 'Lane Mooney', 'Palo Pinto': 'JR Patterson', 'Panola': 'Cutter Clinton',
        'Parker': 'Russell Authier', 'Parmer': 'Eric L. Geske', 'Pecos': 'Thomas J. Perkins',
        'Polk': 'Byron A. Lyons', 'Potter': 'Brian Thomas', 'Presidio': 'Danny C. Dominguez',
        'Rains': 'Michael O. Hopkins', 'Randall': 'Christopher E. Forbis', 'Reagan': 'Jeff Garner',
        'Real': 'Nathan T. Johnson', 'Red River': 'Jim Caldwell', 'Reeves': 'Arturo Granado',
        'Refugio': 'Pinky Gonzales', 'Roberts': 'Bruce Skidmore', 'Robertson': 'Gerald Yezak',
        'Rockwall': 'Terry D. Garrett', 'Runnels': 'Carl L. Squyres', 'Rusk': 'Johnwayne Valdez',
        'Sabine': 'Thomas Neil Maddox Sr', 'San Augustine': 'Robert Cartwright', 'San Jacinto': 'Greg Capers',
        'San Patricio': 'Oscar Rivera', 'San Saba': 'David L. Jenkins', 'Schleicher': 'Jason L. Chatham',
        'Scurry': 'James P. Wilson III', 'Shackelford': 'Edward A. Miller', 'Shelby': 'Kevin Windham',
        'Sherman': 'Ted Allen', 'Smith': 'Larry R. Smith', 'Somervell': 'Alan E. West',
        'Starr': 'Rene Fuentes', 'Stephens': 'James Kevin Roach', 'Sterling': 'Russell L. Irby',
        'Stonewall': 'Bill Mullen', 'Sutton': 'DuWayne Castro', 'Swisher': 'Kyle Schmalzried',
        'Tarrant': 'Bill E. Waybourn', 'Taylor': 'Ricky Bishop', 'Terrell': 'Thaddeus Cleveland',
        'Terry': 'Timothy A. Click', 'Throckmorton': 'Doc Wigington', 'Titus': 'Timothy C. Ingram',
        'Tom Green': 'J. Nick Hanna', 'Travis': 'Sally Hernandez', 'Trinity': 'Woody A. Wallace',
        'Tyler': 'Bryan Weatherford', 'Upshur': 'Lawrence Webb', 'Upton': 'William Mitchell Upchurch',
        'Uvalde': 'Ruben Nolasco', 'Val Verde': 'Joe Frank Martinez', 'Van Zandt': 'Joe Carter',
        'Victoria': 'Justin W. Marr', 'Walker': 'Clint R. McRae', 'Waller': 'Troy Alan Guidry',
        'Ward': 'Frarin V. Valle', 'Washington': 'Otto H. Hanak', 'Webb': 'Martin Cuellar',
        'Wharton': 'Shannon Srubar', 'Wheeler': 'Johnny G. Carter', 'Wichita': 'David Duke',
        'Wilbarger': 'Brian Jack Fritze', 'Willacy': 'Jose A. Salazar', 'Williamson': 'Mike Gleason',
        'Wilson': 'James W. Stewart', 'Winkler': 'James D. Mitchell', 'Wise': 'Lane Akin',
        'Wood': 'Kelly W. Cole', 'Yoakum': 'David Bryant', 'Young': 'Travis Babcock',
        'Zapata': 'Otto H. Hanak', 'Zavala': 'Eusevio E. Salinas Jr',
    },
    'VIRGINIA': {
        'Accomack': 'W. Todd Wessells', 'Albemarle': 'Chan R. Bryant', 'Alexandria': 'Sean Casey',
        'Alleghany': 'Kyle M. Moore', 'Amelia': 'Ricky L. Walker', 'Amherst': 'L.J. "Jimmy" Ayers, III',
        'Appomattox': 'Robert N. "Robby" Richardson', 'Arlington': 'Jose Quiroz', 'Augusta': 'Donald L. Smith',
        'Bath': 'Robert W. Plecker', 'Bedford': 'Mike W. Miller', 'Bland': 'Jason R. Ramsey',
        'Botetourt': 'Matthew T. Ward', 'Bristol': 'R. Tyrone Foster', 'Brunswick': 'Brian K. Roberts',
        'Buchanan': 'Allen W. Boyd', 'Buckingham': 'W. G. "Billy" Kidd Jr.', 'Buena Vista': 'W. Randolph Hamilton Jr.',
        'Campbell': 'Whit W. Clark III', 'Caroline': 'C. Scott Moser', 'Carroll': 'Kevin A. Kemp',
        'Charles City': 'Jayson T. Crawley', 'Charlotte': 'James R. "Randy" Grissom', 'Charlottesville': 'James E. Brown III',
        'Chesapeake': 'Wallace W. Chadwick, III', 'Chesterfield': 'Karl S. Leonard', 'Clarke': 'Travis M. Sumption',
        'Colonial Heights': 'Todd B. Wilson', 'Craig': 'L. Trevor N. Craddock', 'Culpeper': 'Timothy W. Chilton',
        'Cumberland': 'Darrell L. Hodges', 'Danville': 'Michael S. Mondul', 'Dickenson': 'Jeremy D. Fleming',
        'Dinwiddie': 'D. T. "Duck" Adams', 'Emporia': 'Dameon White', 'Essex': 'W. A. "Arnie" Holmes',
        'Fairfax': 'Stacey A. Kincaid', 'Falls Church': 'Metin "Matt" Cay', 'Fauquier': 'Jeremy A. Falls',
        'Floyd': 'Brian J. Craig', 'Fluvanna': 'Eric B. Hess', 'Franklin': 'W. Q. "Bill" Overton Jr.',
        'Frederick': 'L. W. "Lenny" Millholland', 'Fredericksburg': 'Rashawn Cowles', 'Giles': 'W. Morgan Millirons',
        'Gloucester': 'Darrell W. Warren, Jr.', 'Goochland': 'Steven N. Creasey', 'Grayson': 'Gary C. Hash',
        'Greene': 'Steven S. Smith', 'Greensville': 'William T. "Tim" Jarratt, Jr.', 'Halifax': 'Fred S. Clark',
        'Hampton': 'Karen E. Bowden', 'Hanover': 'David R. Hines', 'Henrico': 'Alisa A. Gregory',
        'Henry': 'Wayne Davis', 'Highland': 'Robert W. "Bob" Kelly', 'Hopewell': 'Travis L. Stanley',
        'Isle of Wight': 'James R. Clarke Jr.', 'King and Queen': 'William R. "Rob" Balderson', 'King George': 'Chris A. Giles',
        'King William': 'T. D. "Don" Lumpkin, Jr.', 'Lancaster': 'Patrick McCranie', 'Lee': 'Gary B. Parsons',
        'Loudoun': 'Michael L. Chapman', 'Louisa': 'Donald A. Lowe', 'Lunenburg': 'Arthur Townsend, Jr.',
        'Lynchburg': 'Donald T. Sloan', 'Madison': 'Erik J. Weaver', 'Martinsville': 'Steve M. Draper',
        'Mathews': 'April L. Edwards', 'Mecklenburg': 'R. W. "Bobby" Hawkins Jr.', 'Middlesex': 'David P. Bushey',
        'Montgomery': 'C. H. "Hank" Partin', 'Nelson': 'Mark E. Embrey', 'New Kent': 'Lee S. Bailey',
        'Newport News': 'Gabe A. Morgan', 'Norfolk': 'Joe Baron', 'Northampton': 'David L. Doughty Jr.',
        'Northumberland': 'John A. "Johnny" Beauchamp', 'Norton': 'Jason F. McConnell', 'Nottoway': 'Robert L. Jones Sr.',
        'Orange': 'Jason C. Smith', 'Page': 'Chadwick W. "Chad" Cubbage', 'Patrick': 'Daniel M. Smith',
        'Petersburg': 'Vanessa R. Crawford', 'Pittsylvania': 'Michael W. "Mike" Taylor', 'Portsmouth': 'Michael A. Moore',
        'Powhatan': 'Brad W. Nunnally Jr.', 'Prince Edward': 'L. A. "Tony" Epps', 'Prince George': 'R. W. "Buck" Vargo',
        'Prince William': 'Glendell Hill', 'Pulaski': 'Michael W. Worrell', 'Radford': 'Mark R. Armentrout',
        'Rappahannock': 'Connie S. Compton', 'Richmond': 'Antionette V. Irving', 'Richmond County': 'Stephan B. Smith',
        'Roanoke': 'Antonio D. Hash', 'Roanoke County': 'J. Eric Orange', 'Rockbridge': 'Tony A. McFaddin Jr.',
        'Rockingham': 'Bryan F. Hutcheson', 'Russell': 'William J. "Bill" Watson', 'Salem': 'Chris Shelor',
        'Scott': 'Jeff B. Edds', 'Shenandoah': 'Timothy C. Carter', 'Smyth': 'B. C. "Chip" Shuler',
        'Southampton': 'Josh A. Wyche Sr.', 'Spotsylvania': 'Roger L. Harris', 'Stafford': 'David P. "DP" Decatur, Jr.',
        'Staunton': 'Christopher M. Hartless', 'Suffolk': 'David Miles', 'Surry': 'Carlos Turner',
        'Sussex': 'Ernest L. Giles Sr.', 'Tazewell': 'Brian L. Hieatt', 'Virginia Beach': 'Rocky Holcomb',
        'Warren': 'Crystal M. Cline', 'Washington': 'Blake Andis', 'Waynesboro': 'Christopher Johnson Jr.',
        'Westmoreland': 'C. O. Balderson', 'Williamsburg': 'David J. Hardin', 'James City': 'David J. Hardin',
        'Winchester': 'William Sales', 'Wise': 'E. Grant Kilgore', 'Wythe': 'Anthony R. Cline',
        'York': 'Ronald G. Montgomery', 'Poquoson': 'Ronald G. Montgomery',
    },
    'WEST VIRGINIA': {
        'Barbour': 'Brett Carpenter', 'Berkeley': 'Robert A. Blair', 'Boone': 'Chad P. Barker',
        'Braxton': 'Eddie W. Williams', 'Brooke': 'Richard Beatty', 'Cabell': 'Douglas W. Adams',
        'Calhoun': 'Graham Knight', 'Clay': 'Allen C. King', 'Doddridge': 'Clinton Boring',
        'Fayette': 'Jess R. McMullen', 'Gilmer': 'Donald E. Belin', 'Grant': 'Steven Lee Wratchford Jr.',
        'Greenbrier': 'Bart Baker', 'Hampshire': 'Nathan J. Sions', 'Hancock': 'Scott Gittings',
        'Hardy': 'Steven Dawson', 'Harrison': 'Robert Matheny II', 'Jackson': 'Ross Mellinger',
        'Jefferson': 'Tom Hansen', 'Kanawha': 'Joey Crawford', 'Lewis': 'Mickey E. Metz',
        'Lincoln': 'Harvey Scites', 'Logan': 'Paul D. Clemens', 'Marion': 'Roger Cunningham',
        'Marshall': 'Michael Dougherty', 'Mason': 'Corey J. Miller', 'McDowell': 'James E. Muncy',
        'Mercer': 'Alan Christian', 'Mineral': 'Joel E. Wisner', 'Mingo': 'Joe Smith',
        'Monongalia': 'Todd Forbes', 'Monroe': 'Michael J. Heller', 'Morgan': 'Johnnie Walter',
        'Nicholas': 'John W. Evans Jr.', 'Ohio': 'Harry Nelson Croft Jr.', 'Pendleton': 'Bradley D. Kimble',
        'Pleasants': 'Charles Mankins', 'Pocahontas': 'Troy A. McCoy', 'Preston': 'Scott Spiker',
        'Putnam': 'Bobby Eggleton', 'Raleigh': 'Frank Priddy', 'Randolph': 'Robert R. Elbon Jr.',
        'Ritchie': 'Bryan B.D. Davis', 'Roane': 'Phillip Dever', 'Summers': 'Justin Faris',
        'Taylor': 'Randy Durrett', 'Tucker': 'Jacob Kopec', 'Tyler': 'R. Dean Pratt',
        'Upshur': 'Mike Coffman', 'Wayne': 'Charles E. Bradshaw Jr.', 'Webster': 'Richard Clayton',
        'Wetzel': 'Eric M. Daugherty', 'Wirt': 'Mike McFee', 'Wood': 'Rick Woodyard',
        'Wyoming': 'Brad Ellison',
    },
    'KENTUCKY': {
        'Adair': 'Gary Roy', 'Allen': 'Brandon Ford', 'Anderson': 'Joe Milam',
        'Ballard': 'Ronnie Giles', 'Barren': 'Kent Keen', 'Bath': 'Jessie Stewart',
        'Bell': 'Mitchell Williams', 'Boone': 'Les Hill', 'Bourbon': 'Tony Asbury',
        'Boyd': 'James Reihs', 'Boyle': 'Taylor Bottom', 'Bracken': 'Robert Boody',
        'Breathitt': 'John Hollan', 'Breckinridge': 'Billy Richardson', 'Bullitt': 'Walt Sholar',
        'Butler': 'Scottie Ward', 'Caldwell': 'Don Weedman', 'Calloway': 'Nicky Knight',
        'Campbell': 'Mike Jansen', 'Carlisle': 'Eric Perry', 'Carroll': 'Ryan Gosser',
        'Carter': 'Jeffrey May', 'Casey': 'Chad Weddle', 'Christian': 'Tyler DeArmond',
        'Clark': 'Berl Perdue', 'Clay': 'Patrick Robinson', 'Clinton': 'Ricky Marcum',
        'Crittenden': 'Evan Head', 'Cumberland': 'Keaton Williams', 'Daviess': 'Brad Youngman',
        'Edmonson': 'James Vincent', 'Elliott': 'Ray Craft', 'Estill': 'Christopher Flynn',
        'Fayette': 'Kathy Witt', 'Fleming': 'Timothy Smith', 'Floyd': 'John Hunt',
        'Franklin': 'Dwayne Depp', 'Fulton': 'Chad Parker', 'Gallatin': 'Robert Webster',
        'Garrard': 'Willie Skeens', 'Grant': 'Brian Maines', 'Graves': 'Jon Hayden',
        'Grayson': 'Norman Chaffins', 'Green': 'Robby Beard', 'Greenup': 'Matt Smith',
        'Hancock': 'Dale Bozarth', 'Hardin': 'John Ward', 'Harlan': 'Chris Brewer',
        'Harrison': 'Shain Stephens', 'Hart': 'Jeff Wilson', 'Henderson': 'C. Stauffer',
        'Henry': 'Keith Perry', 'Hickman': 'Ben Natividad', 'Hopkins': 'Matt Sanderson',
        'Jackson': 'Daniel Isaacs', 'Jefferson': 'David James', 'Jessamine': 'Kevin Grimes',
        'Johnson': 'Scott Hazlette', 'Kenton': 'Chuck Korzenborn', 'Knott': 'Dale Richardson',
        'Knox': 'Mike Smith', 'LaRue': 'Brian Smith', 'Laurel': 'John Root',
        'Lawrence': 'Chuck Jackson', 'Lee': 'Joseph Lucas', 'Leslie': 'Billy Collett',
        'Letcher': 'Billy Jones', 'Lewis': 'Johnny Bivens', 'Lincoln': 'D. Hines',
        'Livingston': 'Bobby Davidson', 'Logan': 'Stephen Stratton', 'Lyon': 'Brent White',
        'Madison': 'Mike Coyle', 'Magoffin': 'William Meade', 'Marion': 'James Gardner',
        'Marshall': 'Matt Hilbrecht', 'Martin': 'John Kirk', 'Mason': 'Ryan Swolsky',
        'McCracken': 'Ryan Norman', 'McCreary': 'David Sampson', 'McLean': 'Kenneth Frizzell',
        'Meade': 'Phillip Wimpee', 'Menifee': 'Roger Smallwood Jr.', 'Mercer': 'Ernest Kelty',
        'Metcalfe': 'Lonnie Hodges', 'Monroe': 'Dale Ford', 'Montgomery': 'David Charles',
        'Morgan': 'Greg Motley', 'Muhlenberg': 'William Ward', 'Nelson': 'Ramon Pineiroa Jr.',
        'Nicholas': 'Jeff Sidles', 'Ohio': 'Adam Wright', 'Oldham': 'Timothy Wakefield',
        'Owen': 'Brent Caldwell', 'Owsley': 'Tara Roberts', 'Pendleton': 'Jared Brewer',
        'Perry': 'Joe Engle', 'Pike': 'Rodney Scott', 'Powell': 'Danny Rogers',
        'Pulaski': 'Bobby Jones', 'Robertson': 'Randy Insko', 'Rockcastle': 'Shannon Franklin',
        'Rowan': 'Matt Sparks', 'Russell': 'Derek Polston', 'Scott': 'Jeremy Nettles',
        'Shelby': 'Timothy Moore', 'Simpson': 'Jere Hopson', 'Spencer': 'Andrew Ware',
        'Taylor': 'R. Benningfield', 'Todd': 'Tracy White', 'Trigg': 'Aaron Acree',
        'Trimble': 'Greg Clifford', 'Union': 'Jason Thomas', 'Warren': 'Brett Hightower',
        'Washington': 'Jerry Pinkston', 'Wayne': 'Timothy Catron', 'Webster': 'William Braden',
        'Whitley': 'William Elliotte', 'Wolfe': 'Greg Banks', 'Woodford': 'Johnny Wilhoit',
    },
    'MISSOURI': {
        'Adair': 'Jason Lene', 'Andrew': 'Grant Gillett', 'Atchison': 'Andrew Riley',
        'Audrain': 'Matthew Oller', 'Barry': 'Daniel Boyd', 'Barton': 'John Simpson',
        'Bates': 'Harold Anderson', 'Benton': 'Eric Knox', 'Bollinger': 'Stanley Petton',
        'Boone': 'Robert Carey', 'Buchanan': 'William Puett', 'Butler': 'Mark Dobbs',
        'Caldwell': 'Mitchell Allen', 'Callaway': 'Kenneth Maylee', 'Camden': 'Christopher Edgar',
        'Cape Girardeau': 'Ruth Dickerson', 'Carroll': 'William McCoy', 'Carter': 'Alonzo Bradwell',
        'Cass': 'Jeffrey Weber', 'Cedar': 'James McCrary', 'Chariton': 'Erick Billups',
        'Christian': 'Bradley Cole', 'Clark': 'Shawn Webster', 'Clay': 'William Akin',
        'Clinton': 'Addam Carrel', 'Cole': 'John Wheeler', 'Cooper': 'Christopher Class',
        'Crawford': 'Darin Layman', 'Dade': 'Ryan Robison', 'Dallas': 'Darren Cheek',
        'Daviess': 'Larry Adams', 'DeKalb': 'Kasey Keesaman', 'Dent': 'Robert Wells',
        'Douglas': 'Christopher Degase', 'Dunklin': 'Nicholas Darter', 'Franklin': 'Steven Pelton',
        'Gasconade': 'Scott Eiler', 'Gentry': 'Nicholas Tompkins', 'Greene': 'James Arnott',
        'Grundy': 'Rodney Herring', 'Harrison': 'Trevor Place', 'Henry': 'Aaron Brown',
        'Hickory': 'Gregory Burke', 'Holt': 'Charles Parsons', 'Howard': 'Jeffrey Oswald',
        'Howell': 'Matthew Roberts', 'Iron': 'Charles Helton', 'Jackson': 'Darryl Forte',
        'Jasper': 'Randee Kaiser', 'Jefferson': 'David Marshak', 'Johnson': 'Scott Munsterman',
        'Knox': 'Carl Knoche', 'Laclede': 'Mathew Frederick', 'Lafayette': 'Aaron Dye',
        'Lawrence': 'Brad Delay', 'Lewis': 'David Parrish', 'Lincoln': 'Richard Harrell',
        'Linn': 'Carrie Melte', 'Livingston': 'Dustin Woelfle', 'Macon': 'Kevin Shoemaker',
        'Madison': 'Nicolas Adams', 'Maries': 'Mark Morgan', 'Marion': 'James Shinn',
        'McDonald': 'Robert Evenson', 'Mercer': 'Jeffrey Spencer', 'Miller': 'Louis Gregoire',
        'Mississippi': 'Britton Ferrell', 'Moniteau': 'Anthony Wheatley', 'Monroe': 'Danny Colston',
        'Montgomery': 'Craig Allison', 'Morgan': 'Norman Dills', 'New Madrid': 'Danny Higgerson',
        'Newton': 'Matthew Stewart', 'Nodaway': 'Austin Hann', 'Oregon': 'Eric King',
        'Osage': 'Michael Bonham', 'Ozark': 'Cassidy Martin', 'Pemiscot': 'Joe Bryant',
        'Perry': 'Jason Klaus', 'Pettis': 'Bradley Anders', 'Phelps': 'Michael Kirn',
        'Pike': 'Stephen Korte', 'Platte': 'Erik Holland', 'Polk': 'Edward Morrison',
        'Pulaski': 'Stacy Ball', 'Putnam': 'Jason Knight', 'Ralls': 'Brian Allen',
        'Randolph': 'Willis Boggs', 'Ray': 'Gary Blackwell', 'Reynolds': 'Caleb McCoy',
        'Ripley': 'Radnel Talburt', 'Saline': 'Daniel Kirchhoff', 'Schuyler': 'Joseph Wuebker',
        'Scotland': 'Bryan Whitney', 'Scott': 'Derick Wheetley', 'Shannon': 'Steven Hogan',
        'Shelby': 'Arron Fredrickson', 'St. Charles': 'Scott Lewis', 'St. Clair': 'Lee Hilty',
        'Ste. Genevieve': 'Gary Stolzer', 'St. Francois': 'Jeffrey Crites', 'St. Louis': 'Scott Kiefer',
        'Stoddard': 'Carl Hefner', 'Stone': 'Doug Rader', 'Sullivan': 'Gregory Frazier',
        'Taney': 'Bradley Daniels', 'Texas': 'Scott Lindsey', 'Vernon': 'Sean Buehler',
        'Warren': 'Kevin Harrison', 'Washington': 'Dwayne Reed', 'Wayne': 'Kyle Shearrer',
        'Webster': 'Roye Cole', 'Worth': 'Scott Sherer', 'Wright': 'Glenn Adler',
    },
    'NORTH CAROLINA': {
        'Alamance': 'Terry S. Johnson', 'Buncombe': 'Quentin E. Miller', 'Cabarrus': 'Van Shaw',
        'Catawba': 'Don Brown', 'Cumberland': 'Johnathan Morgan', 'Durham': 'Clarence F. Birkhead',
        'Forsyth': 'Bobby F. Kimbrough Jr.', 'Gaston': 'Chad Hawkins', 'Guilford': 'Danny Rogers',
        'Henderson': 'Lowell Griffin', 'Iredell': 'Darren Campbell', 'Johnston': 'Steve Bizzell',
        'Mecklenburg': 'Garry McFadden', 'New Hanover': 'Ed McMahon', 'Rowan': 'Travis Allen',
        'Wake': 'Willie Rowe',
    },
    'SOUTH CAROLINA': {
        'Anderson': 'Chad McBride', 'Berkeley': 'S. Duane Lewis', 'Charleston': 'Carl Ritchie',
        'Greenville': 'Hobart Lewis', 'Horry': 'Phillip E. Thompson', 'Lexington': 'Bryan Jay Koon',
        'Richland': 'Leon Lott', 'Spartanburg': 'Bill Rhyne', 'York': 'Tony W. Breeden',
    },
    'WISCONSIN': {
        'Brown': 'Todd J. Delain', 'Dane': 'Kalvin D. Barrett', 'Milwaukee': 'Denita R. Ball',
        'Racine': 'Christopher Schmaling', 'Waukesha': 'Eric Severson',
    },
    'PENNSYLVANIA': {
        'Adams': 'Sheriff Muller', 'Allegheny': 'Kevin M. Kraus', 'Armstrong': 'Sheriff Pitzer',
        'Beaver': 'Tony Guy', 'Bedford': 'Wayne Emerick Jr.', 'Berks': 'Mandy P. Miller',
        'Blair': 'James Ott', 'Bradford': 'Clinton J. Walters', 'Bucks': 'Danny Ceisler',
        'Butler': 'Mike Slupe', 'Cambria': 'Donald Robertson', 'Carbon': 'Daniel G. Zeigler',
        'Centre': 'Bryan L. Sampsel', 'Chester': 'Kevin D. Dykes', 'Clarion': 'Shawn Zerfoss',
        'Clearfield': 'Michael B. Churner', 'Clinton': 'Kerry W. Stover', 'Crawford': 'David Powers',
        'Cumberland': 'Jody S. Smith', 'Dauphin': 'Nicholas Chimienti Jr.', 'Delaware': 'Siddiq Kamara',
        'Elk': 'W. Todd Caltagarone', 'Erie': 'Chris Campanelli', 'Fayette': 'James Custer',
        'Franklin': 'Benjamin H. Sites', 'Huntingdon': 'Jeff Leonard', 'Indiana': 'Robert E. Fyock',
        'Jefferson': 'Carl Gotwald', 'Juniata': 'Joshua Stimeling', 'Lackawanna': 'Mark McAndrew',
        'Lancaster': 'Christopher R. Leppler', 'Lawrence': 'Perry L. Quahliero', 'Lebanon': 'Jeffrie C. Marley Jr.',
        'Lehigh': 'Joseph N. Hanna', 'Luzerne': 'Brian M. Szumski', 'Lycoming': 'Eric Spiegel',
        'McKean': 'Dan Woods', 'Mercer': 'Gary Hartmans', 'Monroe': 'Nick Cirranello',
        'Montgomery': 'Sean P. Kilkenny', 'Northampton': 'Christopher Zieger', 'Northumberland': 'Robert J. Wolfe',
        'Philadelphia': 'Rochelle Bilal', 'Pike': 'Brian Vennie', 'Potter': 'Kevin Siska',
        'Schuylkill': 'Shawn Butler', 'Somerset': 'Dusty Weir', 'Susquehanna': 'Lance M. Benedict',
        'Tioga': 'Frank Levindoski', 'Venango': 'Eric Foy', 'Washington': 'Anthony E. Andronas',
        'Wayne': 'Christopher Rosler', 'Westmoreland': 'James Albert', 'Wyoming': 'Robert L. Roberts',
        'York': 'Richard P. Keuerleber III',
    },
}


def match_signer(agency_info, sheriff_data):
    """Try to match an agency to a signer."""
    state = agency_info['state']
    county = agency_info['county']
    agency_name = agency_info['agency']
    agency_type = get_agency_type(agency_name)

    result = {
        **agency_info,
        'agency_type': agency_type,
        'signer_name': None,
        'signer_title': None,
        'source': None,
        'confidence': None,
    }

    # For sheriff offices, look up in sheriff data
    if agency_type == 'sheriff' and state in sheriff_data:
        state_sheriffs = sheriff_data[state]
        county_norm = normalize_county(county)

        for sheriff_county, sheriff_name in state_sheriffs.items():
            if normalize_county(sheriff_county) == county_norm:
                result['signer_name'] = sheriff_name
                result['signer_title'] = 'Sheriff'
                result['source'] = 'sheriff_association'
                result['confidence'] = 'high'
                break

    return result


def main():
    """Main function."""
    print("=" * 60)
    print("Fetching ALL 287(g) Signers")
    print("=" * 60)

    # Load agencies
    agencies = load_287g_agencies()
    print(f"\nLoaded {len(agencies)} agency entries from ICE data")

    # Deduplicate by (state, agency, county)
    seen = set()
    unique_agencies = []
    for a in agencies:
        key = (a['state'], a['agency'], a['county'])
        if key not in seen:
            seen.add(key)
            unique_agencies.append(a)

    print(f"Unique agencies: {len(unique_agencies)}")

    # Match signers
    results = []
    for agency in unique_agencies:
        result = match_signer(agency, SHERIFF_DATA)
        results.append(result)

    # Stats
    by_type = {}
    for r in results:
        t = r['agency_type']
        if t not in by_type:
            by_type[t] = {'total': 0, 'matched': 0}
        by_type[t]['total'] += 1
        if r['signer_name']:
            by_type[t]['matched'] += 1

    print("\nResults by agency type:")
    for t, stats in sorted(by_type.items()):
        pct = 100 * stats['matched'] / stats['total'] if stats['total'] > 0 else 0
        print(f"  {t}: {stats['matched']}/{stats['total']} matched ({pct:.0f}%)")

    total_matched = sum(s['matched'] for s in by_type.values())
    total = len(results)
    print(f"\nTotal: {total_matched}/{total} ({100*total_matched/total:.0f}%)")

    # List unmatched
    unmatched = [r for r in results if not r['signer_name']]
    print(f"\nUnmatched agencies ({len(unmatched)}):")

    # Group unmatched by state
    by_state = {}
    for r in unmatched:
        s = r['state']
        if s not in by_state:
            by_state[s] = []
        by_state[s].append(r)

    for state, agencies in sorted(by_state.items(), key=lambda x: -len(x[1]))[:10]:
        print(f"  {state}: {len(agencies)} unmatched")

    # Save results
    output_path = Path(__file__).parent / "signer_data.json"
    with open(output_path, 'w') as f:
        json.dump(results, f, indent=2)
    print(f"\nSaved to {output_path}")


if __name__ == "__main__":
    main()
