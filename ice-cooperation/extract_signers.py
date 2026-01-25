"""
Extract signer information from 287(g) MOA PDFs.

This script:
1. Extracts MOA URLs from the ICE Excel spreadsheet
2. Downloads the PDFs
3. Parses the signature pages to extract who signed
"""

import json
import re
import os
from pathlib import Path
from urllib.parse import unquote, urlparse, parse_qs

from openpyxl import load_workbook
import requests


def extract_moa_urls(excel_path):
    """Extract MOA URLs from Excel hyperlinks."""
    wb = load_workbook(excel_path)
    ws = wb.active

    moa_data = []

    for row_idx in range(2, ws.max_row + 1):
        state = ws.cell(row=row_idx, column=1).value
        agency = ws.cell(row=row_idx, column=2).value
        agency_type = ws.cell(row=row_idx, column=3).value
        county = ws.cell(row=row_idx, column=4).value
        support_type = ws.cell(row=row_idx, column=5).value
        signed = ws.cell(row=row_idx, column=6).value
        moa_cell = ws.cell(row=row_idx, column=7)

        if not state or not agency:
            continue

        moa_url = None
        if moa_cell.hyperlink:
            url = moa_cell.hyperlink.target

            # Handle Microsoft Safe Links (gcc02.safelinks...)
            if 'safelinks.protection.outlook.com' in url:
                parsed = urlparse(url)
                params = parse_qs(parsed.query)
                if 'url' in params:
                    moa_url = unquote(params['url'][0])
            else:
                moa_url = url

        moa_data.append({
            'state': state.strip() if state else None,
            'agency': agency.strip() if agency else None,
            'agency_type': agency_type.strip() if agency_type else None,
            'county': county.strip() if county else None,
            'support_type': support_type.strip() if support_type else None,
            'signed': signed.strftime('%Y-%m-%d') if signed else None,
            'moa_url': moa_url
        })

    return moa_data


def download_pdf(url, output_path, timeout=30):
    """Download a PDF file."""
    try:
        response = requests.get(url, timeout=timeout)
        if response.status_code == 200:
            with open(output_path, 'wb') as f:
                f.write(response.content)
            return True
    except Exception as e:
        print(f"    Error downloading {url}: {e}")
    return False


def extract_text_from_pdf(pdf_path):
    """Extract text from PDF using pdfplumber or PyPDF2."""
    text = ""

    # Try pdfplumber first
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
        return text
    except ImportError:
        pass
    except Exception as e:
        print(f"    pdfplumber error: {e}")

    # Fall back to PyPDF2
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(pdf_path)
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
        return text
    except ImportError:
        pass
    except Exception as e:
        print(f"    PyPDF2 error: {e}")

    return text


def extract_signature_page(pdf_path):
    """Extract just the signature page text from a PDF."""
    import pdfplumber

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            text = page.extract_text() or ''
            # Signature page contains "By signing this MOA" or "bound thereto"
            if 'By signing this MOA' in text or 'bound thereto' in text:
                return text
    return None


def extract_signer_from_text(text, agency_name):
    """Extract LEA signer name from MOA signature page text.

    The PDF format has two columns: LEA (left) and ICE (right)
    We want the LEA signer (Sheriff, Chief, etc.), not the ICE signer.

    Format example:
        Name· Corey M Shaw Name:- --- ----------
        Title: Chief Of Police Title: Deputy Director
    """
    signer_info = {
        'name': None,
        'title': None
    }

    if not text:
        return signer_info

    # Known ICE signers to exclude (they appear on ICE side)
    ice_signers = ['Madison Sheahan', 'Todd Lyons', 'Todd M. Lyons', 'Enrique M. Lucero',
                   'C.M. Cronen', 'Enrique Lucero']

    # Find signature section
    sig_start = text.find('bound thereto')
    if sig_start < 0:
        sig_start = text.find('For the LEA')
    if sig_start < 0:
        sig_start = 0

    sig_section = text[sig_start:]

    # === Extract title first (Sheriff, Chief of Police, etc.) ===
    title = None
    title_patterns = [
        r'Title[\xb7:\s]+([A-Za-z\s]+?)(?:\s*Title|\s*Agency|\n[A-Z]|\s*Deputy Director|$)',
    ]
    title_match = re.search(title_patterns[0], sig_section)
    if title_match:
        title = title_match.group(1).strip()
        title = re.sub(r'[_\-]+', '', title).strip()
        if title and len(title) > 2 and 'Deputy Director' not in title:
            signer_info['title'] = title

    # === 2025 format: Look for "Name" followed by name (middle dot char 183) ===
    # Pattern: Name followed by · (char 183) or : then a proper name
    name_patterns = [
        # "Name· John Smith" (char 183 = middle dot)
        r'Name[\xb7:\s]+([A-Z][a-z]+(?:\s+[A-Z]\.?)?\s+[A-Z][a-zA-Z]+)',
    ]

    for pattern in name_patterns:
        match = re.search(pattern, sig_section)
        if match:
            name = match.group(1).strip()
            # Exclude ICE signers
            if name not in ice_signers:
                signer_info['name'] = name
                break

    # === 2020 format: Names appear before titles in text ===
    if not signer_info['name']:
        # Look for names followed by LEA-type titles
        lea_title_patterns = [
            (r'([A-Z][a-z]+(?:\s+[A-Z]\.?)?\s+[A-Z][a-zA-Z]+)\s*\n?\s*(?:Deputy )?(?:Sheriff|Chief)', 'Sheriff/Chief'),
            (r'([A-Z][a-z]+(?:\s+[A-Z]\.?)?\s+[A-Z][a-zA-Z]+)\s*\n?\s*Commissioner', 'Commissioner'),
            (r'([A-Z][a-z]+(?:\s+[A-Z]\.?)?\s+[A-Z][a-zA-Z]+)\s*\n?\s*(?:Police )?Director', 'Director'),
            (r'([A-Z][a-z]+(?:\s+[A-Z]\.?)?\s+[A-Z][a-zA-Z]+)\s*\n?\s*Warden', 'Warden'),
        ]
        for pattern, title_type in lea_title_patterns:
            match = re.search(pattern, sig_section)
            if match:
                name = match.group(1).strip()
                if name not in ice_signers:
                    # Verify this isn't from the ICE side by checking context
                    pos = match.start()
                    before = sig_section[max(0, pos-50):pos]
                    if 'ICE' not in before and 'Enforcement and Removal' not in before:
                        signer_info['name'] = name
                        if not signer_info['title']:
                            signer_info['title'] = title_type
                        break

    return signer_info


def main():
    """Main function."""
    print("=" * 60)
    print("287(g) MOA Signer Extraction")
    print("=" * 60)

    excel_path = Path(__file__).parent / "287g_agencies.xlsx"
    moa_dir = Path(__file__).parent / "moas"
    moa_dir.mkdir(exist_ok=True)

    # Extract URLs
    print("\n[1] Extracting MOA URLs from Excel...")
    moa_data = extract_moa_urls(excel_path)
    print(f"    Found {len(moa_data)} entries")

    # Count unique URLs
    unique_urls = set(d['moa_url'] for d in moa_data if d['moa_url'])
    print(f"    Unique MOA URLs: {len(unique_urls)}")

    # For testing, just process first 5
    print("\n[2] Downloading and parsing MOAs (first 10 for testing)...")

    results = []
    processed_urls = set()

    for i, entry in enumerate(moa_data[:50]):  # Limit for testing
        url = entry['moa_url']
        if not url or url in processed_urls:
            continue
        processed_urls.add(url)

        # Generate filename from URL
        filename = url.split('/')[-1]
        if not filename.endswith('.pdf'):
            filename = f"moa_{i}.pdf"

        pdf_path = moa_dir / filename

        print(f"\n  [{i+1}] {entry['state']} - {entry['agency']}")
        print(f"      URL: {url[:80]}...")

        # Download if not exists
        if not pdf_path.exists():
            print(f"      Downloading...")
            if not download_pdf(url, pdf_path):
                print(f"      FAILED to download")
                continue

        # Extract signature page text
        print(f"      Extracting signature page...")
        text = extract_signature_page(pdf_path)

        if not text:
            # Fall back to full text extraction
            text = extract_text_from_pdf(pdf_path)

        if not text:
            print(f"      No text extracted")
            continue

        # Extract signer
        signer = extract_signer_from_text(text, entry['agency'])
        print(f"      Signer: {signer['name']} ({signer['title']})")

        results.append({
            **entry,
            'signer_name': signer['name'],
            'signer_title': signer['title']
        })

    # Save results
    output_path = Path(__file__).parent / "signer_data.json"
    with open(output_path, 'w') as f:
        json.dump(results, f, indent=2)

    print(f"\n[3] Results saved to {output_path}")

    # Print summary
    found = len([r for r in results if r['signer_name']])
    print(f"\n    Successfully extracted {found}/{len(results)} signer names")


if __name__ == "__main__":
    main()
